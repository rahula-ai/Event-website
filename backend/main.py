"""
Dharmasthala Events – FastAPI Backend
======================================
Endpoints:
  POST /api/contact          → Save contact form → MongoDB + Excel → Send email
  GET  /api/contact          → List all submissions (admin)
  GET  /api/contact/export   → Download Excel file of all submissions
  POST /api/newsletter       → Save subscriber + email marketing sync
  GET  /api/newsletter       → List all subscribers (admin)
  POST /api/register         → Event registration → MongoDB + confirmation email
  GET  /api/register         → List all registrations (admin)
  GET  /api/events           → List upcoming events
  GET  /health               → Health check

Auth module (backup – enable with AUTH_ENABLED=true env var):
  POST /api/auth/register
  POST /api/auth/login
  POST /api/auth/logout
  GET  /api/auth/me
"""

import os
import io
import uuid
import smtplib
import logging
import asyncio
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from typing import Optional, List
from contextlib import asynccontextmanager

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from fastapi import FastAPI, HTTPException, Depends, Query, Header, Request, BackgroundTasks
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.trustedhost import TrustedHostMiddleware
from pydantic import BaseModel, EmailStr, Field
from motor.motor_asyncio import AsyncIOMotorClient
import certifi

# ──────────────────────────────────────────────────────────────────────────────
# STRUCTURED LOGGING
# ──────────────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s [%(name)s] %(message)s",
    datefmt="%Y-%m-%dT%H:%M:%S",
)
log = logging.getLogger("dharmasthala")

# ──────────────────────────────────────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────────────────────────────────────
MONGO_URL    = os.getenv("MONGO_URL", "mongodb://localhost:27017")
DB_NAME      = os.getenv("DB_NAME", "dharmasthala_events")
SMTP_HOST    = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT    = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER    = os.getenv("SMTP_USER", "")
SMTP_PASS    = os.getenv("SMTP_PASS", "")
ADMIN_EMAIL  = os.getenv("ADMIN_EMAIL", "admin@dharmasthala.org.in")
FROM_EMAIL   = os.getenv("FROM_EMAIL", "events@dharmasthala.org.in")
ADMIN_KEY    = os.getenv("ADMIN_API_KEY", "change-me-in-production")
ENVIRONMENT  = os.getenv("ENVIRONMENT", "development")

# CORS – comma-separated list of allowed origins (use * only in dev)
_raw_origins    = os.getenv("ALLOWED_ORIGINS", "*")
ALLOWED_ORIGINS = [o.strip() for o in _raw_origins.split(",")] if _raw_origins != "*" else ["*"]

# Newsletter integrations (configure ONE, leave others blank)
MAILCHIMP_API_KEY = os.getenv("MAILCHIMP_API_KEY", "")
MAILCHIMP_LIST_ID = os.getenv("MAILCHIMP_LIST_ID", "")
SENDGRID_API_KEY  = os.getenv("SENDGRID_API_KEY", "")
SENDGRID_LIST_ID  = os.getenv("SENDGRID_LIST_ID", "")

AUTH_ENABLED = os.getenv("AUTH_ENABLED", "false").lower() == "true"

# ──────────────────────────────────────────────────────────────────────────────
# APP SETUP
# ──────────────────────────────────────────────────────────────────────────────
# MongoDB client
client: Optional[AsyncIOMotorClient] = None
db = None


@asynccontextmanager
async def lifespan(application: FastAPI):
    """Manage application lifespan – startup and graceful shutdown."""
    global client, db
    client = AsyncIOMotorClient(MONGO_URL, tlsCAFile=certifi.where())
    db = client[DB_NAME]
    await db.contact_submissions.create_index("created_at")
    await db.newsletter_subscribers.create_index("email", unique=True)
    await db.event_registrations.create_index([("event_id", 1), ("email", 1)])
    log.info("MongoDB connected – db=%s env=%s", DB_NAME, ENVIRONMENT)
    yield
    client.close()
    log.info("MongoDB connection closed")


app = FastAPI(
    title="Dharmasthala Events API",
    description="Backend API for the Dharmasthala Sacred Events website",
    version="1.0.0",
    docs_url="/docs" if ENVIRONMENT != "production" else None,
    redoc_url="/redoc" if ENVIRONMENT != "production" else None,
    lifespan=lifespan,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ──────────────────────────────────────────────────────────────────────────────
# ADMIN AUTH GUARD
# ──────────────────────────────────────────────────────────────────────────────
async def require_admin(x_api_key: str = Header(...)):
    if x_api_key != ADMIN_KEY:
        raise HTTPException(status_code=403, detail="Unauthorized")


# ──────────────────────────────────────────────────────────────────────────────
# PYDANTIC MODELS
# ──────────────────────────────────────────────────────────────────────────────
class ContactForm(BaseModel):
    name: str = Field(..., min_length=2, max_length=100)
    email: EmailStr
    phone: Optional[str] = None
    subject: str = Field(..., min_length=2, max_length=100)
    message: str = Field(..., min_length=10, max_length=2000)

class NewsletterSignup(BaseModel):
    email: EmailStr
    name: Optional[str] = None

class EventRegistration(BaseModel):
    event_id: int
    event_title: str
    name: str = Field(..., min_length=2, max_length=100)
    email: EmailStr
    phone: Optional[str] = None
    guests: int = Field(1, ge=1, le=10)

# ── Auth models (backup module) ──
class UserRegister(BaseModel):
    name: str
    email: EmailStr
    password: str = Field(..., min_length=8)

class UserLogin(BaseModel):
    email: EmailStr
    password: str


# ──────────────────────────────────────────────────────────────────────────────
# EMAIL UTILITIES
# ──────────────────────────────────────────────────────────────────────────────
def _send_email_sync(to: str, subject: str, html_body: str, cc: Optional[str] = None):
    """Blocking SMTP send – always run inside asyncio.to_thread."""
    if not SMTP_USER or not SMTP_PASS:
        log.warning("SMTP not configured – skipping email to %s | subject: %s", to, subject)
        return
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = f"Dharmasthala Events <{FROM_EMAIL}>"
    msg["To"]      = to
    if cc:
        msg["Cc"] = cc
    msg.attach(MIMEText(html_body, "html"))
    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15) as server:
            server.ehlo()
            server.starttls()
            server.ehlo()
            server.login(SMTP_USER, SMTP_PASS)
            recipients = [to] + ([cc] if cc else [])
            server.sendmail(FROM_EMAIL, recipients, msg.as_string())
        log.info("Email sent | to=%s | subject=%s", to, subject)
    except smtplib.SMTPAuthenticationError:
        log.error("SMTP auth failed – check SMTP_USER / SMTP_PASS")
    except smtplib.SMTPException as exc:
        log.error("SMTP error sending to %s: %s", to, exc)
    except Exception as exc:
        log.error("Unexpected email error: %s", exc)


def send_email(to: str, subject: str, html_body: str, cc: Optional[str] = None):
    """Fire-and-forget async wrapper – schedule via background tasks in routes."""
    asyncio.create_task(
        asyncio.to_thread(_send_email_sync, to, subject, html_body, cc)
    )


def contact_admin_email(data: ContactForm) -> str:
    return f"""
    <html><body style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto">
    <div style="background:#38BDF8;padding:24px;border-radius:12px 12px 0 0">
      <h2 style="color:white;margin:0">New Contact Form Submission</h2>
    </div>
    <div style="background:#f8f8f8;padding:24px;border-radius:0 0 12px 12px;border:1px solid #e0e0e0">
      <table style="width:100%;border-collapse:collapse">
        <tr><td style="padding:8px;font-weight:bold;width:120px">Name</td><td style="padding:8px">{data.name}</td></tr>
        <tr style="background:#fff"><td style="padding:8px;font-weight:bold">Email</td><td style="padding:8px"><a href="mailto:{data.email}">{data.email}</a></td></tr>
        <tr><td style="padding:8px;font-weight:bold">Phone</td><td style="padding:8px">{data.phone or '—'}</td></tr>
        <tr style="background:#fff"><td style="padding:8px;font-weight:bold">Subject</td><td style="padding:8px">{data.subject}</td></tr>
        <tr><td style="padding:8px;font-weight:bold;vertical-align:top">Message</td><td style="padding:8px">{data.message}</td></tr>
      </table>
    </div>
    </body></html>
    """


def contact_user_email(data: ContactForm) -> str:
    return f"""
    <html><body style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto">
    <div style="background:linear-gradient(135deg,#38BDF8,#0EA5E9);padding:32px;border-radius:12px 12px 0 0;text-align:center">
      <h1 style="color:white;margin:0;font-size:22px">🙏 Thank You, {data.name}</h1>
    </div>
    <div style="background:#FAFAFA;padding:32px;border-radius:0 0 12px 12px;border:1px solid #e0e0e0">
      <p style="color:#444;font-size:15px">We have received your message and our sevadars will respond within <strong>24–48 hours</strong>.</p>
      <div style="background:#E0F2FE;border-left:4px solid #38BDF8;padding:16px;border-radius:8px;margin:20px 0">
        <strong>Subject:</strong> {data.subject}<br/>
        <strong>Your message:</strong> {data.message}
      </div>
      <p style="color:#666;font-size:13px">May Lord Manjunatha bless your journey.<br/>
      — Dharmasthala Events Team</p>
    </div>
    </body></html>
    """


def registration_confirmation_email(reg: EventRegistration) -> str:
    return f"""
    <html><body style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto">
    <div style="background:linear-gradient(135deg,#38BDF8,#0EA5E9);padding:32px;border-radius:12px 12px 0 0;text-align:center">
      <h1 style="color:white;margin:0">Registration Confirmed! ✅</h1>
    </div>
    <div style="background:#FAFAFA;padding:32px;border-radius:0 0 12px 12px;border:1px solid #e0e0e0">
      <p style="font-size:16px">Dear <strong>{reg.name}</strong>,</p>
      <p>Your registration for <strong>{reg.event_title}</strong> has been confirmed.</p>
      <div style="background:#E0F2FE;border-radius:10px;padding:20px;margin:20px 0">
        <p style="margin:0;font-size:14px"><strong>🎟 Event:</strong> {reg.event_title}<br/>
        <strong>👥 Guests:</strong> {reg.guests} person(s)<br/>
        <strong>📧 Confirmation to:</strong> {reg.email}</p>
      </div>
      <p style="color:#666;font-size:13px">Please arrive at least 30 minutes before the event. 
      Carry this confirmation email as proof of registration.</p>
      <p style="color:#888;font-size:12px">May Lord Manjunatha bless your presence at this sacred event.</p>
    </div>
    </body></html>
    """


# ──────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT UTILITY
# ──────────────────────────────────────────────────────────────────────────────
def make_excel(data: list, sheet_name: str, columns: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header style
    header_fill = PatternFill("solid", fgColor="38BDF8")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="EEEEEE"),
    )

    ws.append([c["label"] for c in columns])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    ws.row_dimensions[1].height = 28

    # Data rows
    for i, row in enumerate(data, start=2):
        ws.append([row.get(c["key"], "") for c in columns])
        for cell in ws[i]:
            cell.alignment = Alignment(vertical="center")
            cell.border = border
        if i % 2 == 0:
            for cell in ws[i]:
                cell.fill = PatternFill("solid", fgColor="F0F9FF")
        ws.row_dimensions[i].height = 22

    # Auto-fit columns
    for i, col in enumerate(columns, start=1):
        max_len = max(len(col["label"]), max((len(str(row.get(col["key"], ""))) for row in data), default=0))
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
# NEWSLETTER INTEGRATION (pluggable)
# ──────────────────────────────────────────────────────────────────────────────
async def sync_newsletter_service(email: str, name: Optional[str]):
    """
    Plug in your preferred email marketing service here.
    Only ONE integration runs at a time based on env vars.
    Future integrations: Brevo, Klaviyo, ConvertKit, etc.
    """
    if MAILCHIMP_API_KEY and MAILCHIMP_LIST_ID:
        try:
            import httpx, base64
            dc = MAILCHIMP_API_KEY.split("-")[-1]
            url = f"https://{dc}.api.mailchimp.com/3.0/lists/{MAILCHIMP_LIST_ID}/members"
            auth = base64.b64encode(f"anystring:{MAILCHIMP_API_KEY}".encode()).decode()
            async with httpx.AsyncClient() as c:
                await c.post(url, json={"email_address": email, "status": "subscribed",
                             "merge_fields": {"FNAME": name or ""}},
                             headers={"Authorization": f"Basic {auth}"}, timeout=10)
            log.info("Mailchimp synced: %s", email)
        except Exception as e:
            log.warning("Mailchimp sync failed: %s", e)

    elif SENDGRID_API_KEY and SENDGRID_LIST_ID:
        try:
            import httpx
            async with httpx.AsyncClient() as c:
                await c.put(
                    f"https://api.sendgrid.com/v3/marketing/contacts",
                    json={"list_ids": [SENDGRID_LIST_ID], "contacts": [{"email": email, "first_name": name or ""}]},
                    headers={"Authorization": f"Bearer {SENDGRID_API_KEY}"}, timeout=10
                )
            log.info("SendGrid synced: %s", email)
        except Exception as e:
            log.warning("SendGrid sync failed: %s", e)

    # Add more integrations below:
    # elif BREVO_API_KEY: ...
    # elif KLAVIYO_API_KEY: ...
    else:
        log.info("Newsletter stored locally (no external service): %s", email)


# ──────────────────────────────────────────────────────────────────────────────
# ROUTES – CONTACT
# ──────────────────────────────────────────────────────────────────────────────
@app.post("/api/contact", status_code=201)
async def submit_contact(form: ContactForm):
    doc = {
        "_id": str(uuid.uuid4()),
        "created_at": datetime.utcnow().isoformat(),
        **form.dict(),
    }
    await db.contact_submissions.insert_one(doc)

    # Send emails asynchronously (fire & forget in production → use background tasks)
    send_email(ADMIN_EMAIL, f"[Contact] {form.subject} – {form.name}", contact_admin_email(form))
    send_email(form.email, "We received your message – Dharmasthala Events", contact_user_email(form))

    return {"success": True, "message": "Thank you! We'll get back to you within 24–48 hours."}


@app.get("/api/contact")
async def list_contacts(_=Depends(require_admin)):
    docs = await db.contact_submissions.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return {"data": docs, "total": len(docs)}


@app.get("/api/contact/export")
async def export_contacts(_=Depends(require_admin)):
    docs = await db.contact_submissions.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    columns = [
        {"label": "Date", "key": "created_at"},
        {"label": "Name", "key": "name"},
        {"label": "Email", "key": "email"},
        {"label": "Phone", "key": "phone"},
        {"label": "Subject", "key": "subject"},
        {"label": "Message", "key": "message"},
    ]
    xlsx = make_excel(docs, "Contact Submissions", columns)
    return StreamingResponse(
        io.BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=contact_submissions.xlsx"}
    )


# ──────────────────────────────────────────────────────────────────────────────
# ROUTES – NEWSLETTER
# ──────────────────────────────────────────────────────────────────────────────
@app.post("/api/newsletter", status_code=201)
async def subscribe_newsletter(data: NewsletterSignup):
    doc = {
        "_id": data.email,
        "email": data.email,
        "name": data.name,
        "subscribed_at": datetime.utcnow().isoformat(),
        "active": True,
    }
    try:
        await db.newsletter_subscribers.insert_one(doc)
    except Exception:
        raise HTTPException(status_code=400, detail="This email is already subscribed.")

    # Sync with email marketing platform
    await sync_newsletter_service(data.email, data.name)

    # Welcome email
    welcome_html = f"""
    <html><body style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto">
    <div style="background:linear-gradient(135deg,#38BDF8,#0EA5E9);padding:32px;text-align:center;border-radius:12px 12px 0 0">
      <h1 style="color:white;margin:0">🙏 Welcome to Our Sacred Community</h1>
    </div>
    <div style="padding:32px;background:#FAFAFA;border:1px solid #e0e0e0;border-radius:0 0 12px 12px">
      <p>Thank you for subscribing to the Dharmasthala Events newsletter!</p>
      <p>You'll receive updates on upcoming festivals, cultural events, and spiritual discourses.</p>
      <p style="color:#888;font-size:12px">To unsubscribe, click <a href="#">here</a>. We respect your privacy.</p>
    </div>
    </body></html>
    """
    send_email(data.email, "Welcome to Dharmasthala Events Newsletter! 🙏", welcome_html)
    return {"success": True, "message": "Subscribed successfully!"}


@app.get("/api/newsletter")
async def list_subscribers(_=Depends(require_admin)):
    docs = await db.newsletter_subscribers.find({}, {"_id": 0}).sort("subscribed_at", -1).to_list(5000)
    return {"data": docs, "total": len(docs)}


@app.get("/api/newsletter/export")
async def export_subscribers(_=Depends(require_admin)):
    docs = await db.newsletter_subscribers.find({}, {"_id": 0}).sort("subscribed_at", -1).to_list(50000)
    columns = [
        {"label": "Email", "key": "email"},
        {"label": "Name", "key": "name"},
        {"label": "Subscribed At", "key": "subscribed_at"},
        {"label": "Active", "key": "active"},
    ]
    xlsx = make_excel(docs, "Newsletter Subscribers", columns)
    return StreamingResponse(
        io.BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=newsletter_subscribers.xlsx"}
    )


# ──────────────────────────────────────────────────────────────────────────────
# ROUTES – EVENT REGISTRATION (backup module, ready to enable)
# ──────────────────────────────────────────────────────────────────────────────
@app.post("/api/register", status_code=201)
async def register_for_event(reg: EventRegistration):
    doc = {
        "_id": str(uuid.uuid4()),
        "created_at": datetime.utcnow().isoformat(),
        **reg.dict(),
    }
    await db.event_registrations.insert_one(doc)

    # Send confirmation email to attendee
    send_email(
        reg.email,
        f"Registration Confirmed – {reg.event_title} | Dharmasthala",
        registration_confirmation_email(reg)
    )
    # Notify admin
    send_email(
        ADMIN_EMAIL,
        f"[New Registration] {reg.event_title} – {reg.name}",
        f"<p><b>{reg.name}</b> ({reg.email}) registered for <b>{reg.event_title}</b> with {reg.guests} guest(s).</p>"
    )
    return {"success": True, "message": "Registration confirmed! Check your email for details."}


@app.get("/api/register")
async def list_registrations(event_id: Optional[int] = Query(None), _=Depends(require_admin)):
    query = {"event_id": event_id} if event_id else {}
    docs = await db.event_registrations.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)
    return {"data": docs, "total": len(docs)}


@app.get("/api/register/export")
async def export_registrations(event_id: Optional[int] = Query(None), _=Depends(require_admin)):
    query = {"event_id": event_id} if event_id else {}
    docs = await db.event_registrations.find(query, {"_id": 0}).sort("created_at", -1).to_list(50000)
    columns = [
        {"label": "Date", "key": "created_at"},
        {"label": "Event", "key": "event_title"},
        {"label": "Name", "key": "name"},
        {"label": "Email", "key": "email"},
        {"label": "Phone", "key": "phone"},
        {"label": "Guests", "key": "guests"},
    ]
    xlsx = make_excel(docs, "Event Registrations", columns)
    return StreamingResponse(
        io.BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=event_registrations.xlsx"}
    )


# ──────────────────────────────────────────────────────────────────────────────
# AUTH MODULE (BACKUP – enabled via AUTH_ENABLED=true)
# ──────────────────────────────────────────────────────────────────────────────
if AUTH_ENABLED:
    import hashlib, secrets

    def hash_password(pw: str, salt: str = "") -> tuple[str, str]:
        if not salt:
            salt = secrets.token_hex(16)
        hashed = hashlib.sha256(f"{salt}{pw}".encode()).hexdigest()
        return hashed, salt

    @app.post("/api/auth/register", status_code=201, tags=["Auth (Backup)"])
    async def auth_register(data: UserRegister):
        existing = await db.users.find_one({"email": data.email})
        if existing:
            raise HTTPException(status_code=400, detail="Email already registered")
        hashed, salt = hash_password(data.password)
        doc = {
            "_id": str(uuid.uuid4()),
            "name": data.name, "email": data.email,
            "password_hash": hashed, "salt": salt,
            "created_at": datetime.utcnow().isoformat(),
            "active": True,
        }
        await db.users.insert_one(doc)
        return {"success": True, "message": "Account created"}

    @app.post("/api/auth/login", tags=["Auth (Backup)"])
    async def auth_login(data: UserLogin):
        user = await db.users.find_one({"email": data.email})
        if not user:
            raise HTTPException(status_code=401, detail="Invalid credentials")
        hashed, _ = hash_password(data.password, user["salt"])
        if hashed != user["password_hash"]:
            raise HTTPException(status_code=401, detail="Invalid credentials")
        token = secrets.token_urlsafe(32)
        await db.sessions.insert_one({
            "_id": token, "user_id": user["_id"],
            "created_at": datetime.utcnow().isoformat()
        })
        return {"token": token, "name": user["name"], "email": user["email"]}

    @app.post("/api/auth/logout", tags=["Auth (Backup)"])
    async def auth_logout(x_auth_token: str = Header(...)):
        await db.sessions.delete_one({"_id": x_auth_token})
        return {"success": True}

    @app.get("/api/auth/me", tags=["Auth (Backup)"])
    async def auth_me(x_auth_token: str = Header(...)):
        session = await db.sessions.find_one({"_id": x_auth_token})
        if not session:
            raise HTTPException(status_code=401, detail="Invalid session")
        user = await db.users.find_one({"_id": session["user_id"]}, {"password_hash": 0, "salt": 0, "_id": 0})
        return user


# ──────────────────────────────────────────────────────────────────────────────
# HEALTH
# ──────────────────────────────────────────────────────────────────────────────
@app.get("/health")
async def health():
    return {
        "status": "ok",
        "auth_enabled": AUTH_ENABLED,
        "smtp_configured": bool(SMTP_USER and SMTP_PASS),
        "newsletter_service": (
            "mailchimp" if MAILCHIMP_API_KEY else
            "sendgrid" if SENDGRID_API_KEY else
            "local_only"
        )
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
